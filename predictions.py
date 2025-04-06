import pandas as pd
import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import webbrowser
import tkinter as tk
from tkinter import messagebox, filedialog
import yfinance as yf
import numpy as np
from sklearn.ensemble import GradientBoostingRegressor
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler
import matplotlib.pyplot as plt
import json
from fpdf import FPDF
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import logging
import os
from datetime import datetime
import plotly.express as px

# Setup logging for error tracking and debugging
logging.basicConfig(filename="app.log", level=logging.DEBUG, format="%(asctime)s - %(levelname)s - %(message)s")

# Sample stock prediction data for Magnificent 7 + Semiconductor stocks
data = {
    "Year": [2025, 2026, 2027, 2028, 2029, 2030],
    "NVIDIA": [0, 60, 120, 180, 240, 300],
    "Apple": [0, 68, 136, 204, 272, 340],
    "Microsoft": [0, 132, 264, 396, 528, 660],
    "Amazon": [0, 64, 128, 192, 256, 320],
    "Meta": [0, 64, 128, 192, 256, 320],
    "Alphabet": [0, 58, 116, 174, 232, 290],
    "Tesla": [0, 80, 160, 240, 320, 400],
    "AMD": [0, 50, 100, 150, 200, 250],
    "Intel": [0, 25, 50, 75, 100, 125],
    "Broadcom": [0, 140, 280, 420, 560, 700],
    "Qualcomm": [0, 48, 96, 144, 192, 240],
    "Texas Instruments": [0, 54, 108, 162, 216, 270]
}

df = pd.DataFrame(data)

# Caching mechanism for fetched stock data
stock_data_cache = {}

# Function to write the DataFrame content to an Excel worksheet
def write_dataframe_to_sheet(dataframe, worksheet):
    """
    Writes the DataFrame into the specified worksheet, placing the data in rows and columns.
    The first row is used for the column headers, followed by the data from the DataFrame.
    """
    for r_idx, row in enumerate(dataframe_to_rows(dataframe, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            worksheet.cell(row=r_idx, column=c_idx, value=value)

# Function to generate Excel with stock forecasts and add charts
def generate_excel_with_charts():
    """
    Generates an Excel file containing the stock forecast data.
    For each stock in the data dictionary, a line chart is created displaying the stock's
    price forecast over the 6 years. The chart is added to the worksheet, which is saved as
    'Tech_Semiconductor_Stock_Forecasts.xlsx'.
    """    # Create a new Excel workbook and add a worksheet
    filename = "Tech_Semiconductor_Stock_Forecasts.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Forecasts"
    write_dataframe_to_sheet(df, ws)

    stocks = list(data.keys())[1:]
    for i, stock in enumerate(stocks):
        chart = LineChart()
        chart.title = f"{stock} 5-Year Stock Price Forecast"
        chart.y_axis.title = "Price (USD)"
        chart.x_axis.title = "Year"

        col = i + 2
        data_ref = Reference(ws, min_col=col, min_row=1, max_row=7)
        cats = Reference(ws, min_col=1, min_row=2, max_row=7)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 10
        chart.width = 20

        col_letter = get_column_letter((i % 5) * 10 + 1)
        row_offset = 2 + (i // 5) * 15
        ws.add_chart(chart, f"{col_letter}{row_offset}")

    wb.save(filename)
    messagebox.showinfo("Success", f"Excel file '{filename}' created with charts.")

# Function to export stock data to a JSON file
def export_to_json():
    """
    Opens a file dialog for the user to select a location to save a JSON file.
    The stock forecast data is saved in JSON format with records for each year.
    """    # Convert DataFrame to JSON format
    file_path = filedialog.asksaveasfilename(defaultextension=".json", filetypes=[("JSON files", "*.json")])
    if file_path:
        df.to_json(file_path, orient="records", indent=4)
        messagebox.showinfo("Success", f"Data exported to {file_path}")

# Function to export stock data to a PDF file
def export_to_pdf():
    """
    Opens a file dialog for the user to select a location to save a PDF file.
    The stock forecast data is written into the PDF in a simple text format.
    """
    file_path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")])
    if file_path:
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=10)
        pdf.cell(200, 10, txt="Stock Forecast Data", ln=True, align='C')
        for index, row in df.iterrows():
            row_text = ", ".join([f"{col}: {row[col]}" for col in df.columns])
            pdf.cell(200, 10, txt=row_text, ln=True)
        pdf.output(file_path)
        messagebox.showinfo("Success", f"PDF exported to {file_path}")

# Function to fetch stock data with caching
def fetch_stock_data(ticker):
    if ticker in stock_data_cache:
        return stock_data_cache[ticker]
    stock = yf.Ticker(ticker)
    data = stock.history(period="5y")
    if data.empty:
        logging.error(f"No data found for ticker: {ticker}")
        return None
    stock_data_cache[ticker] = data
    return data

# Function to preprocess stock data
def preprocess_data(ticker):
    """
    Fetches historical stock data for the given ticker using Yahoo Finance API.
    Prepares the data by extracting features (Open, High, Low, Volume) and labels (Next day's Close price).
    Scales the features using StandardScaler and returns the preprocessed data.
    """
    data = fetch_stock_data(ticker)
    if data is None:
        return None
    data['Tomorrow'] = data['Close'].shift(-1)
    data.dropna(inplace=True)
    X = data[['Open', 'High', 'Low', 'Volume']]
    y = data['Tomorrow']
    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X)
    return X_scaled, y

# Optimized function to open market news related to the selected companies
def open_news():
    """
    Opens a web browser to display market news for the selected stock tickers.
    Users can enter a list of stock tickers or choose 'ALL' to view news for all the stocks.
    The function now handles multiple ticker news requests efficiently.
    """
    # Dictionary of URLs for each stock's news
    urls = {
        "NVIDIA": "https://www.bloomberg.com/news/features/2025-03-14/can-nvidia-stock-go-higher-jensen-huang-looks-to-extend-ai-boom",
        "Apple": "https://www.cnn.com/business/tech/apple",
        "Microsoft": "https://www.bloomberg.com/quote/MSFT:US",
        "Amazon": "https://www.bloomberg.com/quote/AMZN:US",
        "Meta": "https://www.bloomberg.com/quote/META:US",
        "Alphabet": "https://www.bloomberg.com/quote/GOOGL:US",
        "Tesla": "https://www.bloomberg.com/quote/TSLA:US",
        "AMD": "https://www.marketwatch.com/story/amd-is-struggling-against-nvidia-it-could-soon-face-a-resurgent-intel-as-well-8122991c",
        "Intel": "https://www.marketwatch.com/story/amd-is-struggling-against-nvidia-it-could-soon-face-a-resurgent-intel-as-well-8122991c",
        "Broadcom": "https://www.barrons.com/articles/as-nvidia-and-broadcom-fade-3-chip-stocks-for-ais-next-stage-d7a3163d",
        "Qualcomm": "https://www.marketbeat.com/stocks/NASDAQ/NVDA/competitors-and-alternatives/",
        "Texas Instruments": "https://www.marketbeat.com/stocks/NASDAQ/NVDA/competitors-and-alternatives/"
    }

    # Prompt the user to enter tickers or use 'ALL' to see news for all stocks
    choices = tk.simpledialog.askstring("News", "Enter companies (comma separated) or 'ALL' for all:")
    if choices:
        selections = [name.strip() for name in choices.upper().split(",")] if choices != 'ALL' else list(urls.keys())

        # Efficiently open news URLs only for the selected tickers
        for name in selections:
            url = urls.get(name.title())
            if url:
                webbrowser.open(url)
            else:
                messagebox.showwarning("Invalid Ticker", f"No news available for {name}")


# Function to send email alerts
def send_email_alert(subject, body, to_email):
    """
    Sends an email alert to the specified recipient with the provided subject and body content.
    Uses Gmail's SMTP server to send the email. If credentials are not available, prompts the user to enter them.
    """
    from_email = os.getenv("EMAIL_ADDRESS")
    password = os.getenv("EMAIL_PASSWORD")
    if not from_email or not password:
        from_email = tk.simpledialog.askstring("Email", "Enter your email address:")
        password = tk.simpledialog.askstring("Email", "Enter your email password:", show="*")

    msg = MIMEMultipart()
    msg['From'] = from_email
    msg['To'] = to_email
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'plain'))

    try:
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(from_email, password)
        text = msg.as_string()
        server.sendmail(from_email, to_email, text)
        server.quit()
        messagebox.showinfo("Email Sent", f"Alert sent to {to_email}")
    except Exception as e:
        messagebox.showerror("Email Failed", str(e))

# Function to forecast stock prices
def forecast_stock(ticker):
    data = fetch_stock_data(ticker)
    if data is None:
        messagebox.showerror("Error", f"No data found for {ticker}.")
        return None  # Return None if no data is found

    data['Tomorrow'] = data['Close'].shift(-1)
    data.dropna(inplace=True)
    X = data[['Open', 'High', 'Low', 'Volume']]
    y = data['Tomorrow']

    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)
    scaler = StandardScaler()
    X_train = scaler.fit_transform(X_train)
    X_test = scaler.transform(X_test)

    model = GradientBoostingRegressor()
    model.fit(X_train, y_train)

    predicted_price = model.predict([scaler.transform([X.iloc[-1]])[0]])[0]
    accuracy = model.score(X_test, y_test) * 100

    alert_message = f"Predicted Closing Price for Tomorrow ({ticker}): ${predicted_price:.2f}\nModel Accuracy: {accuracy:.2f}%"
    messagebox.showinfo("AI Prediction", alert_message)

    return predicted_price  # Return the predicted price

# Function to compare the AI model accuracy for multiple stock tickers
def compare_tickers(tickers):
    """
    Compares the AI prediction accuracy scores for multiple tickers by fetching data for each ticker,
    training a machine learning model, and displaying the accuracy scores for each stock in a message box.
    """
    scores = {}

    for ticker in tickers:
        try:
            data = fetch_stock_data(ticker)
            if data is None:
                continue
            data['Tomorrow'] = data['Close'].shift(-1)
            data.dropna(inplace=True)
            X = data[['Open', 'High', 'Low', 'Volume']]
            y = data['Tomorrow']

            # Splitting the data into training and testing sets
            X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)
            scaler = StandardScaler()
            X_train = scaler.fit_transform(X_train)
            X_test = scaler.transform(X_test)

            # Training the GradientBoostingRegressor model
            model = GradientBoostingRegressor()
            model.fit(X_train, y_train)

            # Calculating model accuracy
            score = model.score(X_test, y_test) * 100
            scores[ticker] = round(score, 2)
        except Exception as e:
            logging.error(f"Error comparing ticker {ticker}: {e}")
            scores[ticker] = f"Error: {e}"

    # Prepare the message to display the comparison results
    score_text = "\n".join([f"{k}: {v}%" for k, v in scores.items()])

    # Displaying the results in a message box
    messagebox.showinfo("Model Accuracy Scores", score_text)


# GUI setup
root = tk.Tk()
root.title("Tech & Semiconductor Stock Forecast Tool")
root.geometry("550x500")

label = tk.Label(root, text="Tech & Semiconductor Stock Forecast Tool", font=("Arial", 14))
label.pack(pady=10)

# Buttons for the application
button_excel = tk.Button(root, text="Generate Excel Forecasts", command=generate_excel_with_charts)
button_excel.pack(pady=5)

button_json = tk.Button(root, text="Export to JSON", command=export_to_json)
button_json.pack(pady=5)

button_pdf = tk.Button(root, text="Export to PDF", command=export_to_pdf)
button_pdf.pack(pady=5)

button_news = tk.Button(root, text="View Market News", command=open_news)
button_news.pack(pady=5)

entry_label = tk.Label(root, text="Enter Stock Ticker for AI Forecast:")
entry_label.pack()
stock_entry = tk.Entry(root)
stock_entry.pack()

button_forecast = tk.Button(root, text="Forecast with AI", command=lambda: forecast_stock(stock_entry.get().upper()))
button_forecast.pack(pady=5)

entry_label_multi = tk.Label(root, text="Compare Tickers (comma-separated):")
entry_label_multi.pack()
compare_entry = tk.Entry(root)
compare_entry.pack()

button_compare = tk.Button(root, text="Compare AI Scores", command=lambda: compare_tickers(compare_entry.get().upper().split(",")))
button_compare.pack(pady=5)

root.mainloop()
